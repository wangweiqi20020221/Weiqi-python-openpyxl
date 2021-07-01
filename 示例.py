# 导入处理excel的模块openpyxl，此模块可以在命令行中输入pip install openpyxl安装
from openpyxl import load_workbook

def man_or_woman(sheet):
    """
    判断员工是男是女
    如果第三列是0，在第五列生成男
    如果第三列是1，在第五列生成女
    """
    # 遍历整个表
    for index in range(2, 4523):
        # 获取第三列的值
        sex = sheet.cell(row=index, column=3).value
        # 判断sex是0是1
        if sex == 0:
            sheet.cell(row=index, column=5, value="男")
        else:
            sheet.cell(row=index, column=5, value="女")

def count_sex_number(sheet):
    """
    统计男女员工的数量
    """
    # 遍历整个表
    for index in range(2, 4523):
        sex = sheet.cell(row=index, column=3).value
        if sex == 0:
            count["男的数量"] += 1
        else:
            count["女的数量"] += 1
    print(count)

def department_employee_number(employeeSheet, departmentSheet):
    """
    统计各部门员工的数量，并将数据保存在部门表中
    """
    # count用于记录各部门都有多少人
    count = {"营销部": 0, "人事部": 0, "技术部": 0, "财务部": 0, "后勤部": 0}
    # 遍历整个表
    for index in range(2, 4523):
        employeeFromDepartment = employeeSheet.cell(row=index, column=6)
        employeeFromDepartment = employeeFromDepartment.value  # 等同于employeeSheet.cell(row=index, column=6).value
        if employeeFromDepartment == 1:
            count["营销部"] += 1
        elif employeeFromDepartment == 2:
            count["人事部"] += 1
        elif employeeFromDepartment == 3:
            count["技术部"] += 1
        elif employeeFromDepartment == 4:
            count["财务部"] += 1
        elif employeeFromDepartment == 5:
            count["后勤部"] += 1
    # 保存数据到部门表
    departmentSheet.cell(row=2, column=3, value=count["营销部"])
    departmentSheet.cell(row=3, column=3, value=count["人事部"])
    departmentSheet.cell(row=4, column=3, value=count["技术部"])
    departmentSheet.cell(row=5, column=3, value=count["财务部"])
    departmentSheet.cell(row=6, column=3, value=count["后勤部"])

def search_department_leader(employeeSheet, departmentSheet):
    """
    查询部门经理的信息
    """
    # 遍历部门列表
    for index in range(2, 7):
        leaderID = departmentSheet.cell(row=index, column=4).value
        departmentName = departmentSheet.cell(row=index, column=2).value
        # leaderInfo用来存储部门经理的相关信息
        leaderInfo = []
        # 遍历部门经理所在行的所有列
        for j in range(1, 6):
            leaderInfo.append(employeeSheet.cell(row=leaderID+1, column=j).value)
        print("%s部门的领导ID为%d，姓名为%s，性别为%s（%s），年龄为%s" % (departmentName, leaderInfo[0], leaderInfo[1], leaderInfo[2], leaderInfo[4], leaderInfo[3]))

# 程序从这里运行
if __name__ == "__main__":
    # 加载excel文件
    testExcelFile = load_workbook("示例.xlsx")
    # 加载员工表
    employeeSheet = testExcelFile["员工"]
    # 加载部门表
    departmentSheet = testExcelFile["部门"]
    # 调用搜索领导的函数
    search_department_leader(employeeSheet, departmentSheet)
    # 保存excel文件
    testExcelFile.save("示例.xlsx")
    testExcelFile.close()